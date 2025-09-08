#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Actualizador de Fechas de Logística (GUI) para Shopify
- Sube metafields de Orden: Logistica.FechaEntrega (date), Logistica.EnTransito (date)
- Entrada: archivo .xlsx o .csv con columnas: id, FechaEntrega, EnTransito
- GUI con logs, progreso y exporte de errores

Empaquetable con PyInstaller (--onefile --noconsole)

Autor: David Velásquez / ULUM (2025)
"""

import os
import sys
import csv
import time
import threading
import traceback
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from typing import Dict, List, Optional

# --- Dependencias externas mínimas ---
# - requests (necesaria)
# - openpyxl (solo si vas a leer .xlsx)
import requests
try:
    import openpyxl  # para .xlsx
    OPENPYXL = True
except Exception:
    OPENPYXL = False

# ===================== CONFIG SHOPIFY (EDITA ESTO) =====================
SHOPIFY_SHOP_NAME   = "TU-TIENDA.myshopify.com"         # p.ej: "mi-tienda.myshopify.com"
SHOPIFY_ACCESS_TOKEN= "shpua_xxxxxxxxxxxxxxxxxxxxxxxxx" # token Admin API
SHOPIFY_API_VERSION = "2025-01"
# ======================================================================

GRAPHQL_URL = f"https://{SHOPIFY_SHOP_NAME}/admin/api/{SHOPIFY_API_VERSION}/graphql.json"
HEADERS = {"X-Shopify-Access-Token": SHOPIFY_ACCESS_TOKEN, "Content-Type": "application/json"}

# Metafields requeridos
META_FIELDS = [
    {"namespace": "Logistica", "key": "FechaEntrega", "type": "date", "description": "Fecha de entrega confirmada"},
    {"namespace": "Logistica", "key": "EnTransito",  "type": "date", "description": "Fecha de inicio de tránsito"},
]

# Ajustes
BATCH_SIZE = 20
RETRY_ATTEMPTS = 2

# --------------------------- Utilidad GUI ---------------------------
class Logger:
    def __init__(self, text_widget: tk.Text):
        self.text = text_widget

    def write(self, msg: str):
        self.text.configure(state="normal")
        self.text.insert("end", msg)
        self.text.see("end")
        self.text.configure(state="disabled")
        self.text.update_idletasks()

    def info(self, msg: str):
        self.write(f"{msg}\n")

    def sep(self):
        self.write("—" * 60 + "\n")

# --------------------------- Shopify / GraphQL ---------------------------
def graphql_request(url: str, headers: dict, query: str, variables: dict) -> dict:
    resp = requests.post(url, headers=headers, json={"query": query, "variables": variables}, timeout=60)
    resp.raise_for_status()
    data = resp.json()
    if "errors" in data:
        raise RuntimeError(f"GraphQL errors: {data['errors']}")
    return data

def ensure_metafield_definitions(url: str, headers: dict, log: Logger):
    log.info("🔍 Verificando/creando definiciones de metafields…")
    # 1) Intentar leer definición (si existe)
    q_check = """
    query GetDef($ownerType: MetafieldOwnerType!, $ns: String!, $key: String!) {
      metafieldDefinition(ownerType: $ownerType, namespace: $ns, key: $key) { id namespace key }
    }
    """
    # 2) Crear si no existe
    m_create = """
    mutation CreateDef($definition: MetafieldDefinitionInput!) {
      metafieldDefinitionCreate(definition: $definition) {
        createdDefinition { id key namespace }
        userErrors { field message code }
      }
    }
    """
    for mf in META_FIELDS:
        ns, key, mf_type, desc = mf["namespace"], mf["key"], mf["type"], mf["description"]
        try:
            r = graphql_request(url, headers, q_check, {"ownerType": "ORDER", "namespace": ns, "key": key})
            if r.get("data", {}).get("metafieldDefinition"):
                log.info(f"  ✔ Ya existe: {ns}.{key}")
                continue
        except Exception as e:
            log.info(f"  ⚠ Error al verificar {ns}.{key}: {e}. Intentando crear…")

        try:
            payload = {
                "definition": {
                    "name": key, "namespace": ns, "key": key,
                    "description": desc, "type": mf_type, "ownerType": "ORDER"
                }
            }
            r = graphql_request(url, headers, m_create, payload)
            created = r["data"]["metafieldDefinitionCreate"].get("createdDefinition")
            if created:
                log.info(f"  ✔ Creada: {ns}.{key}")
            else:
                errs = r["data"]["metafieldDefinitionCreate"].get("userErrors", [])
                if any((e.get("code") == "TAKEN") or ("already" in (e.get("message","").lower())) for e in errs):
                    log.info(f"  ℹ Ya existente (concurrencia): {ns}.{key}")
                else:
                    log.info(f"  ❌ Error creando {ns}.{key}: {errs}")
        except Exception as e:
            log.info(f"  ❌ Error creando {ns}.{key}: {e}")

def to_gid(order_id: str) -> str:
    try:
        int(order_id)
        return f"gid://shopify/Order/{order_id}"
    except ValueError:
        raise ValueError(f"ID de pedido inválido: {order_id}")

def validate_and_format_date(date_str: str) -> Optional[str]:
    if not date_str:
        return None
    s = str(date_str).strip()
    if not s:
        return None
    # Intentos de parseo comunes → YYYY-MM-DD
    fmts = ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d")
    for fmt in fmts:
        try:
            tm = time.strptime(s, fmt)
            return time.strftime("%Y-%m-%d", tm)
        except ValueError:
            continue
    return None

def load_rows_from_csv(path: str, log: Logger) -> List[Dict[str,str]]:
    rows = []
    log.info(f"📂 Leyendo CSV: {path}")
    with open(path, newline="", encoding="utf-8-sig") as f:
        rdr = csv.DictReader(f)
        for row in rdr:
            r = { (k or "").strip(): (v.strip() if isinstance(v, str) else str(v)) for k,v in row.items() }
            if not r.get("id"):
                continue
            # Campos esperados (opcionalmente vacíos)
            rows.append({"id": r.get("id",""), "FechaEntrega": r.get("FechaEntrega",""), "EnTransito": r.get("EnTransito","")})
    return rows

def load_rows_from_xlsx(path: str, log: Logger) -> List[Dict[str,str]]:
    if not OPENPYXL:
        raise RuntimeError("No se encontró openpyxl. Instálalo antes de leer .xlsx.")
    log.info(f"📂 Leyendo Excel: {path}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    for col in ("id", "FechaEntrega", "EnTransito"):
        if col not in idx:
            raise RuntimeError(f"Columna requerida ausente: {col}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        get = lambda k: "" if row[idx[k]] is None else str(row[idx[k]]).strip()
        _id = get("id")
        if not _id:
            continue
        rows.append({"id": _id, "FechaEntrega": get("FechaEntrega"), "EnTransito": get("EnTransito")})
    wb.close()
    return rows

def read_input_file(path: str, log: Logger) -> List[Dict[str,str]]:
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Archivo no encontrado: {path}")
    if path.lower().endswith(".csv"):
        return load_rows_from_csv(path, log)
    if path.lower().endswith(".xlsx"):
        return load_rows_from_xlsx(path, log)
    raise RuntimeError("Formato no soportado. Usa .csv o .xlsx")

def process_in_batches(url: str, headers: dict, rows: List[Dict[str,str]], log: Logger, progress_cb=None):
    total = len(rows)
    log.info(f"📡 Preparando {total} filas…")
    success, errors = 0, []
    total_batches = (total + BATCH_SIZE - 1) // BATCH_SIZE

    mutation = """
    mutation MetafieldsSet($metafields: [MetafieldsSetInput!]!) {
      metafieldsSet(metafields: $metafields) {
        metafields { ownerId key namespace value }
        userErrors { field message code }
      }
    }
    """

    for bi in range(total_batches):
        start, end = bi * BATCH_SIZE, min((bi+1)*BATCH_SIZE, total)
        batch = rows[start:end]
        metafields_to_set = []
        oids_in_batch = []

        for r in batch:
            order_id = r.get("id","").strip()
            if not order_id:
                continue
            try:
                gid = to_gid(order_id)
                oids_in_batch.append(order_id)
            except ValueError as e:
                errors.append({"order_id": order_id or "<vacío>", "error": str(e)})
                continue

            # FechaEntrega
            v_ent = validate_and_format_date(r.get("FechaEntrega",""))
            if v_ent:
                metafields_to_set.append({
                    "ownerId": gid, "namespace": "Logistica", "key": "FechaEntrega", "type": "date", "value": v_ent
                })
            # EnTransito
            v_trn = validate_and_format_date(r.get("EnTransito",""))
            if v_trn:
                metafields_to_set.append({
                    "ownerId": gid, "namespace": "Logistica", "key": "EnTransito", "type": "date", "value": v_trn
                })

        if not metafields_to_set:
            log.info(f"  ⚠ Lote {bi+1}/{total_batches} sin datos válidos, omitido.")
            if progress_cb: progress_cb(end, total)
            continue

        attempt = 0
        while attempt <= RETRY_ATTEMPTS:
            attempt += 1
            try:
                log.info(f"  🚀 Enviando lote {bi+1}/{total_batches} ({len(oids_in_batch)} órdenes)…")
                r = graphql_request(url, headers, mutation, {"metafields": metafields_to_set})
                uerr = r["data"]["metafieldsSet"].get("userErrors", [])
                if uerr:
                    log.info("    ❗ Errores devueltos por Shopify:")
                    for e in uerr:
                        # atribuimos el mismo error a todas las órdenes del lote (no viene granular)
                        for oid in oids_in_batch:
                            errors.append({"order_id": oid, "error": e.get("message","Error")})
                        log.info(f"      - {e}")
                else:
                    log.info("    ✔ Lote OK")
                    success += len(oids_in_batch)
                break
            except requests.exceptions.RequestException as e:
                if attempt <= RETRY_ATTEMPTS:
                    wait = 2 ** attempt
                    log.info(f"    ⏳ Error de red, reintento en {wait}s… ({e})")
                    time.sleep(wait)
                else:
                    log.info(f"    ❌ Error de red definitivo: {e}")
                    for oid in oids_in_batch:
                        errors.append({"order_id": oid, "error": f"Red: {e}"})
            except Exception as e:
                if attempt <= RETRY_ATTEMPTS:
                    wait = 2 ** attempt
                    log.info(f"    ⏳ Error general, reintento en {wait}s… ({e})")
                    time.sleep(wait)
                else:
                    log.info(f"    ❌ Error general definitivo: {e}")
                    for oid in oids_in_batch:
                        errors.append({"order_id": oid, "error": f"General: {e}"})

        if progress_cb: progress_cb(end, total)

    return success, errors

# --------------------------- GUI ---------------------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Actualizador Fechas Logística - Shopify")
        self.geometry("820x580")
        self.minsize(820, 580)

        self.file_path = tk.StringVar()
        self.failed_rows: List[Dict[str,str]] = []

        # Header
        head = ttk.Frame(self, padding=(10,10))
        head.pack(fill="x")
        ttk.Label(head, text="Archivo (.xlsx o .csv) con columnas: id, FechaEntrega, EnTransito", font=("Segoe UI", 10, "bold")).pack(anchor="w")

        # Selector de archivo
        sel = ttk.Frame(self, padding=(10,5))
        sel.pack(fill="x")
        ttk.Entry(sel, textvariable=self.file_path).pack(side="left", fill="x", expand=True)
        ttk.Button(sel, text="Buscar…", command=self.choose_file).pack(side="left", padx=6)

        # Botones acción
        actions = ttk.Frame(self, padding=(10,5))
        actions.pack(fill="x")
        self.btn_run = ttk.Button(actions, text="Iniciar actualización", command=self.start_process)
        self.btn_run.pack(side="left")
        self.btn_save = ttk.Button(actions, text="Guardar errores CSV", command=self.save_errors, state="disabled")
        self.btn_save.pack(side="left", padx=6)

        # Progreso
        prog = ttk.Frame(self, padding=(10,5))
        prog.pack(fill="x")
        self.progress = ttk.Progressbar(prog, mode="determinate")
        self.progress.pack(fill="x")
        self.lbl_status = ttk.Label(prog, text="Esperando archivo…")
        self.lbl_status.pack(anchor="w", pady=(4,0))

        # Log (Scrolled)
        logf = ttk.Frame(self, padding=(10,10))
        logf.pack(fill="both", expand=True)
        self.text = tk.Text(logf, wrap="word", height=18, state="disabled")
        self.text.pack(side="left", fill="both", expand=True)
        scroll = ttk.Scrollbar(logf, orient="vertical", command=self.text.yview)
        scroll.pack(side="right", fill="y")
        self.text.configure(yscrollcommand=scroll.set)
        self.log = Logger(self.text)

        # Footer
        foot = ttk.Frame(self, padding=(10,6))
        foot.pack(fill="x")
        ttk.Label(foot, text=f"Tienda: {SHOPIFY_SHOP_NAME} | API: {SHOPIFY_API_VERSION}", foreground="#666").pack(anchor="w")

    def choose_file(self):
        path = filedialog.askopenfilename(
            title="Selecciona .xlsx o .csv",
            filetypes=[("Excel/CSV", "*.xlsx *.csv"), ("Excel", "*.xlsx"), ("CSV", "*.csv"), ("Todos", "*.*")]
        )
        if path:
            self.file_path.set(path)
            self.lbl_status.config(text=os.path.basename(path))

    def start_process(self):
        path = self.file_path.get().strip()
        if not path:
            messagebox.showerror("Error", "Selecciona un archivo .xlsx o .csv.")
            return

        self.btn_run.config(state="disabled")
        self.btn_save.config(state="disabled")
        self.progress["value"] = 0
        self.failed_rows = []
        self.text.configure(state="normal"); self.text.delete("1.0", "end"); self.text.configure(state="disabled")
        self.log.info("🔗 Conectando a Shopify…")
        self.log.sep()

        # Correr en hilo aparte para no bloquear la UI
        threading.Thread(target=self._run_pipeline, args=(path,), daemon=True).start()

    def _run_pipeline(self, path: str):
        try:
            # 1) Leer archivo
            try:
                rows = read_input_file(path, self.log)
            except Exception as e:
                self.log.info(f"❌ Error leyendo archivo: {e}")
                self._done_ui(error=True)
                return

            if not rows:
                self.log.info("⚠ No hay filas válidas (se requiere 'id' y alguna fecha).")
                self._done_ui(error=True)
                return

            self.log.info(f"✅ {len(rows)} filas válidas encontradas.")
            self.log.sep()

            # 2) Definiciones de metafields
            try:
                ensure_metafield_definitions(GRAPHQL_URL, HEADERS, self.log)
            except Exception as e:
                self.log.info(f"❌ Error configurando metafields: {e}")
                self._done_ui(error=True)
                return

            # 3) Procesar
            def on_progress(done, total):
                self.progress["maximum"] = total
                self.progress["value"] = done
                self.lbl_status.config(text=f"Progreso: {done}/{total}")

            try:
                ok, errs = process_in_batches(GRAPHQL_URL, HEADERS, rows, self.log, progress_cb=on_progress)
            except Exception as e:
                self.log.info(f"❌ Error crítico durante el procesamiento: {e}")
                self._done_ui(error=True)
                return

            self.log.sep()
            self.log.info("📊 RESULTADO")
            self.log.info(f"   ✔ Exitosas: {ok}")
            self.log.info(f"   ❌ Con errores: {len(errs)}")
            self.log.info(f"   🔢 Total procesadas: {ok + len(errs)}")

            self.failed_rows = errs
            if errs:
                self.btn_save.config(state="normal")
                self.log.info("⚠ Hubo errores. Puedes guardar el detalle con el botón «Guardar errores CSV».")
            else:
                self.log.info("🎉 Todas las órdenes se procesaron correctamente.")
            self._done_ui(error=False)
        except Exception as e:
            # catch-all
            self.log.info("❌ Excepción no controlada:")
            self.log.info(traceback.format_exc())
            self._done_ui(error=True)

    def _done_ui(self, error: bool):
        self.btn_run.config(state="normal")
        if error:
            self.lbl_status.config(text="Finalizado con errores")
        else:
            self.lbl_status.config(text="Proceso finalizado")

    def save_errors(self):
        if not self.failed_rows:
            messagebox.showinfo("Info", "No hay errores que guardar.")
            return
        out = filedialog.asksaveasfilename(
            title="Guardar errores como…",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")]
        )
        if not out:
            return
        try:
            with open(out, "w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=["order_id","error"])
                w.writeheader()
                for row in self.failed_rows:
                    w.writerow(row)
            messagebox.showinfo("Listo", f"Errores guardados en:\n{out}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar: {e}")

def main():
    # Valida credenciales mínimas
    if not SHOPIFY_SHOP_NAME or not SHOPIFY_ACCESS_TOKEN:
        tk.Tk().withdraw()
        messagebox.showerror("Config", "Configura SHOPIFY_SHOP_NAME y SHOPIFY_ACCESS_TOKEN en el script.")
        return
    app = App()
    app.mainloop()

if __name__ == "__main__":
    main()
